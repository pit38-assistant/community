#!/usr/bin/env python3

# /// script
# requires-python = ">=3.12"
# dependencies = [
#     "openpyxl",
# ]
# ///

import argparse
import csv
import os
from collections import defaultdict
from datetime import datetime
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook


BROKER = 'SAXO'

# Polish country names -> ISO 2-letter codes (Saxo uses Polish locale)
_COUNTRY_MAP = {
    'Arabia Saudyjska': 'SA',
    'Australia': 'AU',
    'Austria': 'AT',
    'Belgia': 'BE',
    'Brazylia': 'BR',
    'Bułgaria': 'BG',
    'Chile': 'CL',
    'Chiny': 'CN',
    'Chorwacja': 'HR',
    'Cypr': 'CY',
    'Czechy': 'CZ',
    'Dania': 'DK',
    'Egipt': 'EG',
    'Estonia': 'EE',
    'Filipiny': 'PH',
    'Finlandia': 'FI',
    'Francja': 'FR',
    'Grecja': 'GR',
    'Hiszpania': 'ES',
    'Holandia': 'NL',
    'Hongkong': 'HK',
    'Indie': 'IN',
    'Indonezja': 'ID',
    'Irlandia': 'IE',
    'Islandia': 'IS',
    'Izrael': 'IL',
    'Japonia': 'JP',
    'Kanada': 'CA',
    'Katar': 'QA',
    'Korea Południowa': 'KR',
    'Litwa': 'LT',
    'Luksemburg': 'LU',
    'Łotwa': 'LV',
    'Malezja': 'MY',
    'Malta': 'MT',
    'Meksyk': 'MX',
    'Niderlandy': 'NL',
    'Niemcy': 'DE',
    'Norwegia': 'NO',
    'Nowa Zelandia': 'NZ',
    'Portugalia': 'PT',
    'RPA': 'ZA',
    'Republika Południowej Afryki': 'ZA',
    'Rumunia': 'RO',
    'Singapur': 'SG',
    'Słowacja': 'SK',
    'Słowenia': 'SI',
    'Stany Zjednoczone': 'US',
    'Szwajcaria': 'CH',
    'Szwecja': 'SE',
    'Tajwan': 'TW',
    'Turcja': 'TR',
    'Ukraina': 'UA',
    'Węgry': 'HU',
    'Wielka Brytania': 'GB',
    'Wietnam': 'VN',
    'Włochy': 'IT',
    'Zjednoczone Emiraty Arabskie': 'AE',
}


def convert_saxo(workbook_path: str) -> tuple[list[dict], list[dict]]:
    """Convert Saxo Bank XLSX tax report.

    Reads sheets: 'PNL', 'Trading Costs', 'Revenues', 'WithHoldings'.
    Returns (trade_rows, income_rows).
    """
    wb = load_workbook(workbook_path, read_only=True, data_only=True)

    costs = _load_trading_costs(wb['Trading Costs'])
    wht_map = _load_withholdings(wb['WithHoldings'])

    trade_rows = _process_pnl(wb['PNL'], costs)
    income_rows = _process_revenues(wb['Revenues'], wht_map)

    wb.close()
    return trade_rows, income_rows


def _iter_dicts(sheet):
    rows = sheet.iter_rows(values_only=True)
    headers = [str(h).strip() for h in next(rows)]
    for values in rows:
        if not any(values):
            continue
        yield dict(zip(headers, values))


def _parse_amount(value) -> Decimal:
    """Parse numeric or Polish-formatted string ('2130,96') -> Decimal."""
    if value is None:
        return Decimal('0')
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    s = str(value).strip().replace(',', '.')
    return Decimal(s) if s else Decimal('0')


def _parse_date(value) -> str:
    """Parse '20250924' -> '2025-09-24'."""
    return datetime.strptime(str(value).strip(), '%Y%m%d').date().isoformat()


def _fix_mojibake(s: str) -> str:
    """Saxo stores UTF-8 bytes as cp1252 chars (e.g. 'WÅ‚ochy' instead of 'Włochy')."""
    try:
        return s.encode('cp1252').decode('utf-8')
    except (UnicodeEncodeError, UnicodeDecodeError):
        return s


def _country_iso(name: str) -> str:
    name = _fix_mojibake(name.strip())
    return _COUNTRY_MAP.get(name, name.replace(' ', ''))


def _load_trading_costs(sheet) -> dict:
    """Returns {trade_id_str: Decimal(abs_commission)}."""
    costs = defaultdict(Decimal)
    for row in _iter_dicts(sheet):
        trade_id = str(row.get('Trade Id', '')).strip()
        amount = _parse_amount(row.get('Amount'))
        if trade_id:
            costs[trade_id] += abs(amount)
    return dict(costs)


def _load_withholdings(sheet) -> dict:
    """Returns {corporate_action_id_str: (Decimal(abs_wht), currency)}."""
    wht = {}
    for row in _iter_dicts(sheet):
        ca_id = str(row.get('Corporate Action ID', '')).strip()
        amount = abs(_parse_amount(row.get('Amount')))
        currency = str(row.get('Currency Code', '')).strip()
        if ca_id:
            prev_amount, _ = wht.get(ca_id, (Decimal('0'), currency))
            wht[ca_id] = (prev_amount + amount, currency)
    return wht


def _process_pnl(sheet, costs: dict) -> list[dict]:
    # Saxo's PNL sheet is a FIFO match report: one trade split across multiple
    # lot-pairs appears as multiple rows sharing the same Buy/Sell Trade Id.
    # Aggregate quantity and amount per (direction, tx_id) to reconstruct real trades.
    # Value Date is the sell's settlement; buy settlement isn't reported, so we
    # fall back to Buy Trade Date.
    merged: dict[tuple[str, str], dict] = {}

    for row in _iter_dicts(sheet):
        symbol_code = str(row.get('Instrument Symbol Code', '')).strip()
        symbol = symbol_code.split(':')[0] if ':' in symbol_code else symbol_code
        country = _country_iso(str(row.get('Issuer country Name', '')))
        currency = str(row.get('Currency Code', '')).strip()
        quantity = _parse_amount(row.get('Settled Quantity'))
        sell_trade_date = _parse_date(row.get('Sell Trade Date'))
        sell_settlement = _parse_date(row.get('Value Date'))
        buy_trade_date = _parse_date(row.get('Buy Trade Date'))

        for direction, id_col, price_col, value_col, trade_date, settlement in (
            ('SELL', 'Sell Trade Id', 'Sell Price', 'Value of Sell',
             sell_trade_date, sell_settlement),
            ('BUY', 'Buy Trade Id', 'Buy Price', 'Value of Buy',
             buy_trade_date, buy_trade_date),
        ):
            tx_id = str(row.get(id_col, '')).strip()
            amount = _parse_amount(row.get(value_col))
            existing = merged.get((direction, tx_id))
            if existing:
                existing['quantity'] += quantity
                existing['amount'] += amount
            else:
                merged[(direction, tx_id)] = {
                    'broker': BROKER,
                    'tx_id': tx_id,
                    'direction': direction,
                    'symbol': symbol,
                    'isin': '',
                    'country': country,
                    'currency': currency,
                    'price': _parse_amount(row.get(price_col)),
                    'quantity': quantity,
                    'amount': amount,
                    'commission': costs.get(tx_id, Decimal('0')),
                    'operation_datetime': trade_date,
                    'settlement_date': settlement,
                }

    rows = []
    for d in merged.values():
        for k in ('price', 'quantity', 'amount', 'commission'):
            d[k] = str(d[k])
        rows.append(d)
    return rows


def _process_revenues(sheet, wht_map: dict) -> list[dict]:
    rows = []
    for row in _iter_dicts(sheet):
        bk_type = str(row.get('BK Amount Type', '')).strip()
        if 'Dividend' not in bk_type and 'Cash' not in bk_type:
            continue

        ca_id = str(row.get('Corporate Action ID', '')).strip()
        bk_amount_id = str(row.get('Bk Amount Id', '')).strip()
        currency = str(row.get('Currency Code', '')).strip()
        value_date = _parse_date(row.get('Value Date'))
        gross_amount = _parse_amount(row.get('Amount'))

        wht_amount, _ = wht_map.get(ca_id, (Decimal('0'), currency))
        issuer_country = _country_iso(str(row.get('Issuer Country Name', '')))

        rows.append({
            'broker': BROKER,
            'tx_id': bk_amount_id,
            'income_type': 'DIVIDEND',
            'symbol': f'{issuer_country}-{currency}-DIV',
            'isin': '',
            'country': issuer_country,
            'currency': currency,
            'gross_amount': str(gross_amount),
            'wht_amount': str(wht_amount),
            'operation_datetime': value_date,
            'settlement_date': value_date,
        })

    return rows


def _write_csv(outfile, rows: list[dict]) -> None:
    writer = csv.DictWriter(outfile, fieldnames=list(rows[0].keys()))
    writer.writeheader()
    writer.writerows(rows)


def main():
    parser = argparse.ArgumentParser(description='Convert Saxo Bank XLSX tax report to CSV')
    parser.add_argument('input', help='Input Saxo Bank .xlsx file')
    parser.add_argument('--trades-output', help='Output trades CSV')
    parser.add_argument('--income-output', help='Output income CSV')

    args = parser.parse_args()

    trade_rows, income_rows = convert_saxo(args.input)
    stem = Path(args.input).stem
    suffix = os.urandom(3).hex()

    if trade_rows:
        trades_path = args.trades_output or f'result_trades_{stem}_{suffix}.csv'
        with open(trades_path, 'w') as f:
            _write_csv(f, trade_rows)
        print(f"Wrote {len(trade_rows)} trade(s) → {trades_path}")

    if income_rows:
        income_path = args.income_output or f'result_income_{stem}_{suffix}.csv'
        with open(income_path, 'w') as f:
            _write_csv(f, income_rows)
        print(f"Wrote {len(income_rows)} income record(s) → {income_path}")


if __name__ == '__main__':
    main()
